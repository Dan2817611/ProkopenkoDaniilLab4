import csv
import openpyxl
from datetime import datetime

def calculate_employer_age(DateBirth):
    today = datetime.today()
    employer_age = today.year - DateBirth.year - ((today.month, today.day) < (DateBirth.month, DateBirth.day))
    return employer_age

try:
    with open('file_exp1.csv', mode='r', newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)

        workbook = openpyxl.Workbook()

        if 'Sheet' in workbook.sheetnames:
            sheet = workbook['Sheet']
            workbook.remove(sheet)

        worksheet_all = workbook.create_sheet(title="all")
        worksheet_all.append(["Прізвище", "Ім’я", "По батькові", "Дата народження", "Стать", "Посада", "Місто проживання", "Адреса проживання", "Телефон", "Email"])

        worksheet_younger_18 = workbook.create_sheet(title="younger_18")
        worksheet_younger_18.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        worksheet_18_45 = workbook.create_sheet(title="18-45")
        worksheet_18_45.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        worksheet_45_70 = workbook.create_sheet(title="45-70")
        worksheet_45_70.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        worksheet_older_70 = workbook.create_sheet(title="older_70")
        worksheet_older_70.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        count_all = 0
        count_younger_18 = 0
        count_18_45 = 0
        count_45_70 = 0
        count_older_70 = 0

        for row in csv_reader:
            birthdate = datetime.strptime(row["Дата народження"], '%d.%m.%Y')
            age = calculate_employer_age(birthdate)

            if age < 18:
                category = "younger_18"
                count_younger_18 += 1
                row_number = count_younger_18
            elif 18 <= age <= 45:
                category = "18-45"
                count_18_45 += 1
                row_number = count_18_45
            elif 45 < age <= 70:
                category = "45-70"
                count_45_70 += 1
                row_number = count_45_70
            else:
                category = "older_70"
                count_older_70 += 1
                row_number = count_older_70

            count_all += 1
            worksheet_all.append([row["Прізвище"], row["Ім’я"], row["По батькові"], row["Дата народження"], row["Стать"], row["Посада"], row["Місто проживання"], row["Адреса проживання"], row["Телефон"], row["Email"]])

            worksheet_category = workbook[category]
            worksheet_category.append([row_number, row["Прізвище"], row["Ім’я"], row["По батькові"], row["Дата народження"], age])

        workbook.save("file_exp2.xlsx")
        print("Файл file_exp2.xlsx створено")
except FileNotFoundError:
    print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV.")
except Exception as e:
    print("Повідомлення про неможливість створення XLSX файлу:", str(e))
